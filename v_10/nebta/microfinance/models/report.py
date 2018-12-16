# -*- coding: utf-8 -*-
# Author : Mudathir Ahmed
# Part of Odoo. See LICENSE file for full copyright and licensing details.


from odoo import api, fields, models, tools
import datetime
from odoo.report.report_sxw import report_sxw
from odoo.api import Environment
from cStringIO import StringIO
from dateutil.relativedelta import relativedelta
from datetime import datetime
import openpyxl


class finance_report_funding_gender(models.Model):
    _name = "finance.report.funding.genderreport"
    _description = "Funding Gender Report"
    _auto = False

    amount = fields.Float(string='Funded Amount', readonly=True, group_operator='sum')  # group_operator='avg'
    company_id = fields.Many2one('res.company', string="Branch", default=lambda self: self.env.user.company_id,
                                 readonly=True, required=True, ondelete='cascade')
    formula = fields.Selection([('fixed_murabaha', 'Fixed Murabaha'), ('dec_murabaha', 'Decremental Murabaha'),
                                ('salam', 'Salam'), ('ejara', 'Ejara'), ('gard_hassan', 'Gard Hassan'),
                                ('estisnaa', 'Estisnaa'), ('mugawla', 'Mugawla'), ('mudarba', 'Mudarba'),
                                ('musharka', 'Musharka'), ('muzaraa', 'Muzaraa')], string='Formula',
                               required=True)

    asl = fields.Float('Assest Amount')
    rebih = fields.Float('Profit Amount')
    #installmentstate = fields.Selection(
    #    [('draft', 'Draft'), ('delay', 'Delay'), ('adverse', 'Adverse'), ('done', 'Done')]
    #    , string="Installemnt State")



    order_id = fields.Char('Order ID')
    #indiv_id = fields.Char('Indiv ID')
    type = fields.Selection([('individual', 'Individual Funding'), ('group', 'Group Funding')],string="Order Type")
    #num_all=fields.Integer('Gender Num')
    gender = fields.Selection([('male', 'Male'), ('female', 'Female')],string="Gender")
    user_id= fields.Many2one('res.users', string="Officer", readonly=True, required=True,ondelete='restrict')  # restrict
    sector_id = fields.Many2one('finance.sector', string='Sector', ondelete='restrict', required=True)
    portfolio_id = fields.Many2one('finance.portfolio', string="Portfolio")
    #customer = fields.Char('res.partner', string='Customer', ondelete='restrict')
    customer = fields.Char(string='Customer')
    customer_count = fields.Integer(string='Customer Count')
    #installmentstate = fields.Selection([('draft', 'Draft'), ('delay', 'Delay'), ('adverse', 'Adverse'), ('done', 'Done')],
    #                          string='State')
    installment_due_date = fields.Date(string="Installment Date")
    approve_amount_order=fields.Float('Approve Amount')
    approve_amount_order = fields.Float('Approve Amount Order')
    order_state_together = fields.Selection([('a', 'Recieved Requests'), ('b', 'Files Visited'),('c', 'Approved'),('d', 'Funded')],string="Order State Together")
    order_date = fields.Date(string="Order Date")
    visit_date = fields.Date(string="Visit Date")
    amount_residual_inst=fields.Float('Amount Residual')
    amount_residual_inst_state=fields.Char('Residual State')




    @api.model_cr
    def init(self):
        tools.drop_view_if_exists(self._cr, 'finance_report_funding_genderreport')
        self._cr.execute(

            """ CREATE OR REPLACE VIEW finance_report_funding_genderreport AS (
            select * from ((select dd.order_id as id,dd.order_id as order_id ,dd.indiv_id as indiv_id,dd.type ,dd.num_all,

 (case when dd.gender is null or  dd.gender not like 'female' and num_all >0 then 'male' else 'female' end ) as gender,
 dd.appr_idd,dd.approve_amount_order as approve_amount_order,dd.customer_count,
 dd.asl,
 dd.rebih,
 dd.asl+dd.rebih as amount,
 dd.installment_due_date,
 dd.ml,
 dd.amount_residual,
 dd.amount_residual_inst,
 dd.amount_residual_inst_state,
 dd.company_id,dd.user_id,
 dd.formula,dd.sector_id,
 dd.portfolio_id,
 dd.order_state,dd.order_state_together,
 dd.order_date,dd.visit_date,
 dd.customer
 from (
 select ord.id as order_id,
 iord.id as indiv_id,
 ord.type,
(case when prt.gender like 'female' or prt.gender like 'male'
      and ord.type like 'individual'  then 1
      when ord.type like 'group' then gord.male 
end) as num_all,prt.gender,appr.id as appr_idd,
(case when ord.type='individual' then appr.approve_amount/ ((select count(ordd.id) from finance_order ordd  full join finance_group_order gord on ordd.id=gord.order_id
		
    			full join finance_individual_order iord on ordd.id=iord.order_id
    			full join finance_visit vis on ordd.id=vis.order_id
    			full join finance_approval apprr on vis.id=apprr.visit_id
    			full join finance_installments inst on apprr.id = inst.approval_id
    			full join res_company comp on ordd.company_id=comp.id
    			left join finance_portfolio portf on ordd.portfolio_id = portf.id
    			left join res_partner prt on ordd.partner_id=prt.id
    			left join finance_sector sect on ordd.sector_id = sect.id
    			join res_users usr on ordd.user_id=usr.id where apprr.id = appr.id  ))/2

    			
      when ord.type='group' then (appr.approve_amount/((select sum(gord.male+gord.female) from finance_order ordd  full join finance_group_order gord on ord.id=gord.order_id
		
    			full join finance_individual_order iord on ordd.id=iord.order_id
    			full join finance_visit vis on ordd.id=vis.order_id
    			full join finance_approval apprr on vis.id=appr.visit_id
    			full join finance_installments inst on apprr.id = inst.approval_id
    			full join res_company comp on ord.company_id=comp.id
    			left join finance_portfolio portf on ord.portfolio_id = portf.id
    			left join res_partner prt on ord.partner_id=prt.id
    			left join finance_sector sect on ordd.sector_id = sect.id
    			join res_users usr on ordd.user_id=usr.id where apprr.id = appr.id  )))*gord.male  end) as approve_amount_order
,
(case when ord.type='individual' then sum(inst.amount_before_profit )/2
      when ord.type='group' then (sum(inst.amount_before_profit )/(gord.male+gord.female))*gord.male  end) as "asl"
      ,(case when ord.type='individual' then sum(inst.profit_amount)/2
      when ord.type='group' then (sum(inst.profit_amount)/(gord.male+gord.female))*gord.male  end) as "rebih"
      
      ,inst.due_date as installment_due_date,ml.id as ml,ml.amount_residual as amount_residual,
      (case when ord.type='individual' then sum(ml.amount_residual)/2
	    when ord.type='group' then (sum(ml.amount_residual)/(gord.male+gord.female))*gord.male  end)  as amount_residual_inst,
	    (CASE
	    WHEN ml.amount_residual > 0 AND inst.due_date < CURRENT_DATE then 'adverse'
	    WHEN ml.amount_residual > 0 AND inst.due_date < CURRENT_DATE then 'delay'
            WHEN ml.amount_residual = 0 then 'done'
            ELSE 'draft' 
            END)
            as amount_residual_inst_state,
      ((case
      when ord.type = 'individual' then 1/2::float 
      when ord.type = 'group' then (gord.male) end)::float/ 
      (select (case when count(dd.typee)=0 then 1 else count(dd.typee) end) from
      
(select ord.id as order_id ,ord.type as typee  from finance_order ord
 join finance_visit vis on vis.order_id=ord.id
 join finance_approval appr on appr.visit_id = vis.id
full join finance_installments inst on inst.approval_id=appr.id  
 order by ord.id) as dd where dd.order_id= ord.id and dd.typee=  ord.type)
)::float
 as customer_count
 ,ord.company_id
,ord.user_id,appr.formula,ord.sector_id,ord.portfolio_id,prt.name as "customer",
ord.state as order_state, (case when ord.state = 'draft' or ord.state='waiting_visit' then 'a'
				when ord.state = 'visit_complete' or ord.state = 'su_recommend' or ord.state = 'br_recommend' or ord.state = 'op_recommend' then 'b' 
				when ord.state = 'approved' and inst.profit_amount  is Null then 'c'
				when inst.profit_amount > 0 or inst.amount_before_profit > 0 then 'd'
				end) as order_state_together
				,ord.date as order_date,vis.date as visit_date
from finance_order ord  full join finance_group_order gord on ord.id=gord.order_id
		
    			full join finance_individual_order iord on ord.id=iord.order_id
    			full join finance_visit vis on ord.id=vis.order_id
    			full join finance_approval appr on vis.id=appr.visit_id
    			full join finance_installments inst on appr.id = inst.approval_id
    			full join res_company comp on ord.company_id=comp.id
    			left join finance_portfolio portf on ord.portfolio_id = portf.id
    			left join res_partner prt on ord.partner_id=prt.id
    			left join finance_sector sect on ord.sector_id = sect.id
    			join res_users usr on ord.user_id=usr.id
    			left join account_move_line ml ON inst.line_id=ml.id 

    	group by iord.id, ord.id ,vis.id,comp.name,portf.name,
    	inst.due_date,prt.name,sect.name,gord.male,
    	appr.formula,appr.id,usr.login,inst.id,appr.id,gord.female,prt.gender,
    	appr.approve_amount,customer_count,ml.id
                 order by ord.id)  as dd)



                 

                 union all




(select dd.order_id as id ,dd.order_id as order_id ,dd.indiv_id,dd.type ,dd.num_all,
 (case when dd.gender is null or  dd.gender not like 'male' and num_all >0 then 'female' else 'male' end ) as gender,
 dd.appr_idd,dd.approve_amount as approve_amount_order,dd.customer_count,
 dd.asl,dd.rebih,dd.asl+dd.rebih as amount,
 dd.installment_due_date,dd.ml,dd.amount_residual,dd.amount_residual_inst,dd.amount_residual_inst_state
 ,dd.company_id,dd.user_id,dd.formula,dd.sector_id,dd.portfolio_id
 ,dd.order_state
 ,dd.order_state_together,dd.order_date,dd.visit_date,dd.customer
from (select ord.id as order_id,iord.id as indiv_id,ord.type,appr.id as appr_idd,appr.approve_amount as approve_amount_order,

(case when prt.gender like 'female' or prt.gender like 'male' and
 ord.type like 'individual'  then 1 when ord.type like 'group' then gord.female 
end) as num_all,prt.gender,
appr.id as appr_id,
(case when ord.type='individual' then appr.approve_amount/ ((select count(ordd.id) from finance_order ordd  full join finance_group_order gord on ord.id=gord.order_id
		
    			full join finance_individual_order iord on ordd.id=iord.order_id
    			full join finance_visit vis on ordd.id=vis.order_id
    			full join finance_approval apprr on vis.id=appr.visit_id
    			full join finance_installments inst on apprr.id = inst.approval_id
    			full join res_company comp on ord.company_id=comp.id
    			left join finance_portfolio portf on ord.portfolio_id = portf.id
    			left join res_partner prt on ord.partner_id=prt.id
    			left join finance_sector sect on ordd.sector_id = sect.id
    			join res_users usr on ordd.user_id=usr.id where apprr.id = appr.id  ))/2
      when ord.type='group' then (appr.approve_amount/((select sum(gord.male+gord.female) from finance_order ordd  full join finance_group_order gord on ord.id=gord.order_id
		
    			full join finance_individual_order iord on ordd.id=iord.order_id
    			full join finance_visit vis on ordd.id=vis.order_id
    			full join finance_approval apprr on vis.id=appr.visit_id
    			full join finance_installments inst on apprr.id = inst.approval_id
    			full join res_company comp on ord.company_id=comp.id
    			left join finance_portfolio portf on ord.portfolio_id = portf.id
    			left join res_partner prt on ord.partner_id=prt.id
    			left join finance_sector sect on ordd.sector_id = sect.id
    			join res_users usr on ordd.user_id=usr.id where apprr.id = appr.id  )))*gord.female  end)as approve_amount
,
(case when ord.type='individual' then sum(inst.amount_before_profit )/2
      when ord.type='group' then (sum(inst.amount_before_profit )/(gord.male+gord.female))*gord.female  end) as "asl"
      ,(case when ord.type='individual' then sum(inst.profit_amount)/2
      when ord.type='group' then (sum(inst.profit_amount)/(gord.male+gord.female))*gord.female  end) as "rebih"
    
      ,inst.due_date as installment_due_date,
      (case when ord.type='individual' then sum(ml.amount_residual)/2
	    when ord.type='group' then (sum(ml.amount_residual)/(gord.male+gord.female))*gord.female  end)  as amount_residual_inst,
	    (CASE
	    WHEN ml.amount_residual > 0 AND inst.due_date < CURRENT_DATE then 'adverse'
	    WHEN ml.amount_residual > 0 AND inst.due_date < CURRENT_DATE then 'delay'
            WHEN ml.amount_residual = 0 then 'done'
            ELSE 'draft' 
            END)
            as amount_residual_inst_state,
      ((case
      when ord.type = 'individual' then 1/2::float
      when ord.type = 'group' then (gord.female) end)::float/ 
      (select (case when count(dd.typee)=0 then 1 else count(dd.typee) end) from
(select ord.id as order_id ,ord.type as typee  from finance_order ord
 join finance_visit vis on vis.order_id=ord.id
 join finance_approval appr on appr.visit_id = vis.id
full join finance_installments inst on inst.approval_id=appr.id  
 order by ord.id) as dd where dd.order_id= ord.id and dd.typee=  ord.type)
)::float
 as customer_count
 , ord.company_id
 , ord.user_id
 ,appr.formula,ord.sector_id,ord.portfolio_id,prt.name as "customer"
 ,ord.state as order_state
 ,(case when ord.state = 'draft' or ord.state='waiting_visit' then 'a'
				when ord.state = 'visit_complete' or ord.state = 'su_recommend' or ord.state = 'br_recommend' or ord.state = 'op_recommend' then 'b' 
				when ord.state = 'approved' and inst.profit_amount  is Null then 'c'
				when inst.profit_amount > 0 or inst.amount_before_profit > 0 then 'd'
				end) as order_state_together,ord.date as order_date,vis.date as visit_date,ml.id as ml,ml.amount_residual as amount_residual
from finance_order ord  full join finance_group_order gord on ord.id=gord.order_id
		
    			full join finance_individual_order iord on ord.id=iord.order_id
    			full join finance_visit vis on ord.id=vis.order_id
    			full join finance_approval appr on vis.id=appr.visit_id
    			full join finance_installments inst on appr.id = inst.approval_id
    			full join res_company comp on ord.company_id=comp.id
    			left join finance_portfolio portf on ord.portfolio_id = portf.id
    			left join res_partner prt on ord.partner_id=prt.id
    			left join finance_sector sect on ord.sector_id = sect.id
    			join res_users usr on ord.user_id=usr.id 
    			left join account_move_line ml ON inst.line_id=ml.id

    	group by iord.id,ord.id ,vis.id,comp.name,portf.name,
    	inst.due_date,prt.name,sect.name,gord.male,
    	appr.formula,appr.id,usr.login,inst.id,appr.id,
    	gord.female,prt.gender,appr.approve_amount,customer_count,ml.id
                 order by ord.id)  as dd)) as dd order by dd.id
            
            )
            """)


class finance_report_portfolio_profit(models.Model):
    _name = "finance.report.portfolio.profit"
    _description = "Portfolio Profit"
    _auto = False
    port_name = fields.Char(string='Portfolio Name')
    #profit need to be correct to calculate portfolio profit #dont forget
    profit = fields.Float(string='Amount', readonly=True, group_operator='sum')  # group_operator='avg'
    date = fields.Date(string='Date')
    branch = fields.Char('Branch')
    portfolio_funding_amount = fields.Float('Funding Amount')

    @api.model_cr
    def init(self):
        tools.drop_view_if_exists(self._cr, 'finance_report_portfolio_profit')
        self._cr.execute( """ CREATE OR REPLACE VIEW finance_report_portfolio_profit AS (
            
            select port.id as id ,port.name as port_name,comp.name as branch 
            ,sum(appr.approve_amount) As "portfolio_funding_amount"
             
            from 
            finance_portfolio port join
            finance_order ord on ord.portfolio_id = port.id join
            res_company comp on ord.company_id = comp.id join 
            finance_visit vis on ord.visit_id=vis.id join
            finance_approval appr on appr.visit_id=vis.id 

                        group by port.id,port.name,comp.name,ord.date
                            )


            """)


############################## SEEP ##############################


class account_financial_report(models.Model):
    _inherit = 'account.financial.report'

    line = fields.Integer(string='Line', help="Line in seep tmplate")
    select = fields.Selection([('debit', 'Debit'), ('credit', 'Credit'),
                               ('balance', 'Balance')],
                              help="if this for seep then select from what you want accounts to calculate")



class finance_report_seep_template_wizard(models.TransientModel):
    _name = "finance.report.seep.template.wizard"
    _description = "Seep Template Wizard"

    main_year = fields.Integer('Year for Current Data', size=4,required = 1 )
    first_month_year = fields.Integer('First Month Of Financial Year', size=2, default=1,required = 1)
    first_year = fields.Selection(
        [('1', 'Enter Monthly Data'), ('2', 'Enter Quarterly Data'), ('3', 'Enter Annual Data')],required = 1)
    second_year = fields.Selection(
        [('1', 'Enter Monthly Data'), ('2', 'Enter Quarterly Data'), ('3', 'Enter Annual Data'), ('4', 'OFF')],required = 1)
    third_year = fields.Selection(
        [('1', 'Enter Monthly Data'), ('2', 'Enter Quarterly Data'), ('3', 'Enter Annual Data'), ('4', 'OFF')])
    forth_year = fields.Selection([('1', 'Enter Annual Data'), ('2', 'OFF')])
    fifth_year = fields.Selection([('1', 'Enter Annual Data'), ('2', 'OFF')])





    def save_to_template(self):

        # excel file path start from odoo main directory
        file = 'odooseep.xlsx'

        wb = openpyxl.load_workbook(filename=file, keep_vba=True)
        # Selet Sheet Bu Name
        ws = wb.get_sheet_by_name('Sheet1')
        # Insert Basic Year Data To Correct Position in SEEP Excel

        ws['D23'] = self.main_year
        ws['D24'] = self.first_month_year

        ws['D27'] = self.fifth_year
        ws['D28'] = self.forth_year
        ws['D29'] = self.third_year
        ws['D30'] = self.second_year
        ws['D31'] = self.first_year

        #in case upper year is off then make all years below it off in seep template
        if (self.second_year == '4'):
            #self.third_year = 4
            ws['D29'] = '4'
            #self.forth_year = 2
            ws['D28'] = '2'
            #self.fifth_year = 2
            ws['D27'] = '2'
        elif (self.third_year == '4'):
            #self.forth_year = 2
            ws['D28'] = '2'
            #self.fifth_year = 2
            ws['D27'] = '2'
        elif (self.forth_year == '2'):
            #self.fifth_year = 2
            ws['D27'] = '2'


        # to set positions in excel columns
        # Save all required position in tuples
        col_posi = []
        year = []
        main_cells_date = []
        main_start_date = datetime.strptime(str(self.main_year + 1) + '-' + str(self.first_month_year) + '-1',
                                            "%Y-%m-%d")
        months = relativedelta(months=self.first_month_year)
        years = relativedelta(years=self.main_year)


        if (self.fifth_year == '1'):
            year.append(self.main_year - 4)
            col_posi.append('K')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-12) + relativedelta(years=-4))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(years=-4)))

        if (self.forth_year == '1'):
            year.append(self.main_year - 3)
            col_posi.append('M')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-12) + relativedelta(years=-3))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(years=-3)))

        if (self.third_year == '1'):
            year.append(self.main_year - 2)
            col_posi.append('O')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-12) + relativedelta(years=-2))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-11) + relativedelta(years=-2)))

            col_posi.append('P')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-11) + relativedelta(years=-2))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-10) + relativedelta(years=-2)))

            col_posi.append('Q')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-10) + relativedelta(years=-2))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-9) + relativedelta(years=-2)))

            col_posi.append('R')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-9) + relativedelta(years=-2))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-8) + relativedelta(years=-2)))

            col_posi.append('S')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-8) + relativedelta(years=-2))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-7) + relativedelta(years=-2)))

            col_posi.append('T')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-7) + relativedelta(years=-2))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-6) + relativedelta(years=-2)))

            col_posi.append('U')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-6) + relativedelta(years=-2))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-5) + relativedelta(years=-2)))

            col_posi.append('V')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-5) + relativedelta(years=-2))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-4) + relativedelta(years=-2)))

            col_posi.append('W')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-4) + relativedelta(years=-2))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-3) + relativedelta(years=-2)))

            col_posi.append('X')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-3) + relativedelta(years=-2))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-2) + relativedelta(years=-2)))

            col_posi.append('Y')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-2) + relativedelta(years=-2))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-1) + relativedelta(years=-2)))

            col_posi.append('Z')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-1) + relativedelta(years=-2))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-0) + relativedelta(years=-2)))


        elif (self.third_year == '2'):
            year.append(self.main_year - 2)
            col_posi.append('Q')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-12) + relativedelta(years=-2))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-9) + relativedelta(years=-2)))

            col_posi.append('T')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-9) + relativedelta(years=-2))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-6) + relativedelta(years=-2)))

            col_posi.append('W')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-6) + relativedelta(years=-2))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-3) + relativedelta(years=-2)))

            col_posi.append('Z')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-3) + relativedelta(years=-2))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-0) + relativedelta(years=-2)))

        elif (self.third_year == '3'):
            year.append(self.main_year - 2)
            col_posi.append('Z')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-12) + relativedelta(years=-2))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(years=-2)))

        if (self.second_year == '1'):
            year.append(self.main_year - 1)
            col_posi.append('AB')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-12) + relativedelta(years=-1))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-11) + relativedelta(years=-1)))

            col_posi.append('AC')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-11) + relativedelta(years=-1))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-10) + relativedelta(years=-1)))

            col_posi.append('AD')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-10) + relativedelta(years=-1))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-9) + relativedelta(years=-1)))

            col_posi.append('AE')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-9) + relativedelta(years=-1))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-8) + relativedelta(years=-1)))

            col_posi.append('AF')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-8) + relativedelta(years=-1))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-7) + relativedelta(years=-1)))

            col_posi.append('AG')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-7) + relativedelta(years=-1))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-6) + relativedelta(years=-1)))

            col_posi.append('AH')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-6) + relativedelta(years=-1))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-5) + relativedelta(years=-1)))

            col_posi.append('AI')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-5) + relativedelta(years=-1))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-4) + relativedelta(years=-1)))

            col_posi.append('AG')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-4) + relativedelta(years=-1))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-3) + relativedelta(years=-1)))

            col_posi.append('AK')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-3) + relativedelta(years=-1))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-2) + relativedelta(years=-1)))

            col_posi.append('AL')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-2) + relativedelta(years=-1))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-1) + relativedelta(years=-1)))

            col_posi.append('AM')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-1) + relativedelta(years=-1))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-0) + relativedelta(years=-1)))

        elif (self.second_year == '2'):
            year.append(self.main_year - 1)
            col_posi.append('AB')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-12) + relativedelta(years=-1))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-9) + relativedelta(years=-1)))

            col_posi.append('AG')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-9) + relativedelta(years=-1))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-6) + relativedelta(years=-1)))

            col_posi.append('AJ')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-6) + relativedelta(years=-1))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-3) + relativedelta(years=-1)))

            col_posi.append('AM')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-3) + relativedelta(years=-1))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-0) + relativedelta(years=-1)))

        elif (self.second_year == '3'):
            year.append(self.main_year - 1)
            col_posi.append('AM')
            main_cells_date.append('between ' + str(
                (main_start_date + relativedelta(months=-12) + relativedelta(years=-1))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(years=-1)))

        if (self.first_year == '1'):
            year.append(self.main_year)
            col_posi.append('AO')
            main_cells_date.append('between ' + str((main_start_date + relativedelta(months=-12))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-11)))

            col_posi.append('AP')
            main_cells_date.append('between ' + str((main_start_date + relativedelta(months=-11))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-10)))

            col_posi.append('AQ')
            main_cells_date.append('between ' + str((main_start_date + relativedelta(months=-10))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-9)))

            col_posi.append('AR')
            main_cells_date.append('between ' + str((main_start_date + relativedelta(months=-9))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-8)))

            col_posi.append('AS')
            main_cells_date.append('between ' + str((main_start_date + relativedelta(months=-8))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-7)))

            col_posi.append('AT')
            main_cells_date.append('between ' + str((main_start_date + relativedelta(months=-7))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-6)))

            col_posi.append('AU')
            main_cells_date.append('between ' + str((main_start_date + relativedelta(months=-6))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-5)))

            col_posi.append('AV')
            main_cells_date.append('between ' + str((main_start_date + relativedelta(months=-5))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-4)))

            col_posi.append('AW')
            main_cells_date.append('between ' + str((main_start_date + relativedelta(months=-4))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-3)))

            col_posi.append('AX')
            main_cells_date.append('between ' + str((main_start_date + relativedelta(months=-3))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-2)))

            col_posi.append('AY')
            main_cells_date.append('between ' + str((main_start_date + relativedelta(months=-2))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-1)))

            col_posi.append('AZ')
            main_cells_date.append('between ' + str((main_start_date + relativedelta(months=-1))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-0)))

        elif (self.first_year == '2'):
            year.append(self.main_year)
            col_posi.append('AQ')

            main_cells_date.append('between ' + str((main_start_date + relativedelta(months=-12))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-9)))

            col_posi.append('AT')

            main_cells_date.append('between ' + str((main_start_date + relativedelta(months=-9))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-6)))

            col_posi.append('AW')

            main_cells_date.append('between ' + str((main_start_date + relativedelta(months=-6))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-3)))

            col_posi.append('AZ')

            main_cells_date.append('between ' + str((main_start_date + relativedelta(months=-3))) + ' and ' + str(
                main_start_date + relativedelta(days=-1) + relativedelta(months=-0)))

        elif (self.first_year == '3'):
            year.append(self.main_year)
            col_posi.append('AZ')

            main_cells_date.append('between ' + str((main_start_date + relativedelta(months=-12))) + ' and ' + str(
                main_start_date + relativedelta(days=-1)))

        main_start_date = datetime.strptime(str(self.main_year) + '-' + str(self.first_month_year) + '-1', "%Y-%m-%d")




        ############################### Section 1 : Calculate Accounting #################################

        #get all accounts in accounts.financial.report under name SEEP and calculate depit and credit and balance and
        #  save line in seep template

        query = '''SELECT  account_id as id, SUM(debit) AS debit, SUM(credit) AS credit, (SUM(debit) - SUM(credit)) AS balance
                          FROM  account_move_line  ml join account_account acc on acc.id = ml.account_id
                 where account_id in  and ml.date between %s and %s
                     GROUP BY account_id'''
        sum = 0
        account = []
        accounts = []
        query_result = 0
        i = 0 # to change date index
        for doc in self.env['account.financial.report'].search([('parent_id.name', '=', 'SEEP')]):
            if (doc.parent_id.name == 'SEEP'):
                for acc in doc.account_ids:

                    accounts.append(acc.id)
                accounts_tuple = str(tuple(accounts))
                #print ">>>>>>>>>>>>>>>>>>>>>>>>",accounts_tuple[:-2]+')' ,accounts, len(accounts)

                if len(accounts) == 1 :
                    query = '''SELECT  account_id AS id, SUM(debit) AS debit, SUM(credit) AS credit, (SUM(debit) - SUM(credit)) AS balance
                                                                  FROM  account_move_line  ml join account_account acc on acc.id = ml.account_id
                                                         where account_id in ''' + str(tuple(accounts))[:-2]+')' + ''' and ml.date between %s and %s
                                                             GROUP BY account_id'''
                else:
                    query = '''SELECT  account_id AS id, SUM(debit) AS debit, SUM(credit) AS credit, (SUM(debit) - SUM(credit)) AS balance
                                                                                      FROM  account_move_line  ml join account_account acc on acc.id = ml.account_id
                                                                             where account_id in ''' + str(tuple(accounts)) + ''' and ml.date between %s and %s
                                                                                 GROUP BY account_id'''
                for col_pos in col_posi:
                    #query = '''SELECT  account_id AS id, SUM(debit) AS debit, SUM(credit) AS credit, (SUM(debit) - SUM(credit)) AS balance
                    #                          FROM  account_move_line  ml join account_account acc on acc.id = ml.account_id
                    #                 where account_id in '''+ accounts_tuple +''' and ml.date between %s and %s
                    #                     GROUP BY account_id'''

                    self._cr.execute(query,[main_cells_date[i][8:18], main_cells_date[i][32:42]])
                    result = self._cr.fetchall()
                    if result:
                        if(doc.select == 'depit'):
                            query_result = result[0][1]
                        elif(doc.select == 'credit'):
                            query_result = result[0][2]
                        elif (doc.select == 'balance'):
                            query_result = result[0][3]
                    else:
                        query_result = 0
                    #print ">>>>>>>>>>>>>>>>>>>>>>>>>>query_result ",query_result, col_pos + str(doc.line)
                    ws[col_pos + str(doc.line)] = query_result  # Number of Loans Disbursed
                    #if(query_result == 5000.0):
                    #    print ">>>>>>>>>>>>>>>>>>>>>>>>col_pos ",col_pos , doc.line , col_pos + str(doc.line)
                    i += 1
                i = 0
                accounts=[]
        accounts = []


        ###################### Section 2 : Calculate Operation Data ####################

        # Start of calculation
        i = 0
        # Portfolio Data
        """
        240 - 
        """

        # Number of Loans Disbursed (YTD)
        # line 240
        # Value of Loans Disbursed (YTD)
        # line 241


        query = '''select count(id) , sum(amount_before_profit)
                 Loan_Value from finance_installments where due_Date between
                  %s and %s
                        '''

        for col_pos in col_posi:
            self._cr.execute(query, [main_cells_date[i][8:18], main_cells_date[i][32:42]])
            result = self._cr.fetchall()
            query_result = result[0][0]
            if not query_result:
                query_result = 0

            query_result2 = result[0][1]
            if not query_result2:
                query_result2 = 0
            ws[col_pos + str(240)] = query_result  # Number of Loans Disbursed
            ws[col_pos + str(241)] = query_result2  # Value of Loans Disbursed
            i += 1
        i = 0

        # Number of Loans Outstanding (EOP)
        # line 242
        # Value of Loans Outstanding (EOP)
        # line 249


        query = ''' select count(i.id) ,
                 sum(amount_before_profit)
                 Loan_Value from finance_installments i
                 inner join
                 account_move_line ml on i.line_id=ml.id 
                 where
                 ml.amount_residual > 0 and due_Date   < %s
                 '''

        for col_pos in col_posi:
            self._cr.execute(query, [main_cells_date[i][32:42]])
            result = self._cr.fetchall()
            query_result = result[0][0]
            if not query_result:
                query_result = 0

            query_result2 = result[0][1]
            if not query_result2:
                query_result2 = 0
            ws[col_pos + str(242)] = query_result  # Number of Loans Disbursed
            ws[col_pos + str(249)] = query_result2  # Value of Loans Disbursed
            i += 1
        i = 0



        # Numbers Of Loans in Current Portfolio
        # line 273
        #Value Of Loans in Current Portfolio
        # line289

        query = ''' select count(i.id) ,
                         sum(amount_before_profit)
                         Loan_Value from finance_installments i
                         inner join
                         account_move_line ml on i.line_id=ml.id 
                         where
                         ml.amount_residual > 0 and due_Date   >= %s '''

        for col_pos in col_posi:
            self._cr.execute(query, [main_cells_date[i][32:42]])
            result = self._cr.fetchall()
            query_result = result[0][0]
            if not query_result:
                query_result = 0

            query_result2 = result[0][1]
            if not query_result2:
                query_result2 = 0
            ws[col_pos + str(273)] = query_result  # Number of Loans in current portfolio
            ws[col_pos + str(289)] = query_result2  # Value of Loans in current portfolio

            i += 1
        i = 0

        # Numbers Loans at Risk 31 to 90 days
        # line 274
        # Value Loans at Risk 31 to 90 days
        # line 289

        query = ''' select count(i.id) ,
                                 sum(amount_before_profit)
                                 Loan_Value from finance_installments i
                                 inner join
                                 account_move_line ml on i.line_id=ml.id 
                                 where
                                 ml.amount_residual > 0 and due_Date + INTERVAL '90 day'   < %s '''

        for col_pos in col_posi:
            self._cr.execute(query, [main_cells_date[i][32:42]])
            result = self._cr.fetchall()
            query_result = result[0][0]
            if not query_result:
                query_result = 0

            query_result2 = result[0][1]
            if not query_result2:
                query_result2 = 0
            ws[col_pos + str(273)] = query_result  # Number of Loans in current portfolio
            ws[col_pos + str(289)] = query_result2  # Value of Loans in current portfolio

            i += 1
        i = 0

        # Numbers Loans at Risk 180 days
        # line 275
        # Value Loans at Risk 180 days
        # line 290

        query = ''' select count(i.id) ,
                                         sum(amount_before_profit)
                                         Loan_Value from finance_installments i
                                         inner join
                                         account_move_line ml on i.line_id=ml.id 
                                         where
                                         ml.amount_residual > 0  and due_Date + INTERVAL '180 day'   < %s   '''

        for col_pos in col_posi:
            self._cr.execute(query, [main_cells_date[i][32:42]])
            result = self._cr.fetchall()
            query_result = result[0][0]
            if not query_result:
                query_result = 0

            query_result2 = result[0][1]
            if not query_result2:
                query_result2 = 0
            ws[col_pos + str(273)] = query_result  # Number of Loans in current portfolio
            ws[col_pos + str(289)] = query_result2  # Value of Loans in current portfolio

            i += 1
        i = 0



        # Non-Financial Data Stat HERE!!
        # LINES
        """
        383 - 390
        """

        # number of active client
        query = '''
                select count(partner_id),sum(clients) from (
        		select ord.partner_id , sum((case when ord.type = 'group' then gord.female+gord.male when ord.type ='individual' then 1 else 0 end)) clients from
        		finance_installments i
        		inner join
        		account_move_line ml on i.line_id=ml.id
        		join finance_approval app on i.approval_id=app.id
        		join finance_visit vis on app.visit_id = vis.id
        		join finance_order ord on vis.order_id = ord.id
        		full join finance_individual_order indiv on indiv.order_id = ord.id 
			full join finance_group_order gord on gord.order_id = ord.id 

        		where i.due_date < %s
        		group by ord.partner_id
        		having sum(ml.amount_residual) > 0 ) s'''

        for col_pos in col_posi:
            self._cr.execute(query, [main_cells_date[i][32:42]])
            query_result = self._cr.fetchall()[0][1]

            ws[col_pos + str(383)] = query_result
            ws[col_pos + str(385)] = query_result  # Number of Active Borrowers
            i += 1
        i = 0

        # Number of New Clients during period

        query = '''select sum((case when ord.type = 'group' then gord.female+gord.male when ord.type ='individual' then 1 else 0 end)) clients 
                    from finance_order ord 
                    full join finance_individual_order indiv on indiv.order_id = ord.id 
                    full join finance_group_order gord on gord.order_id = ord.id 
                    where ord.date  between %s and %s '''


        for col_pos in col_posi:
            #self._cr.execute(query, [main_cells_date[i][8:18], main_cells_date[i][32:42]])
            ids = []
            sum_new_clients=0
            for line in self.env['finance.approval.payment'].search([('date', '>=', main_cells_date[i][8:18]) , ('date', '<=', main_cells_date[i][32:42])]):
                ids.append(line.approval_id.visit_id.order_id.id)

            for line in self.env['finance.individual.order'].search([('order_id','in',ids)]):
                sum_new_clients += 1
            for line in self.env['finance.group.order'].search([('order_id','in',ids)]):
                sum_new_clients += line.male + line.female

            #query_result = self._cr.fetchall()[0][0]
            query_result = sum_new_clients
            if not query_result:
                query_result = 0

            ws[col_pos + str(384)] = query_result
            i += 1
        i = 0

        # Number of Active Borrowers during period
        # line 385

        query = '''select count(partner_id),sum(clients) from (
        		select ord.partner_id , sum((case when ord.type = 'group' then gord.female+gord.male when ord.type ='individual' then 1 else 0 end)) clients from
        		finance_installments i
        		inner join
        		account_move_line ml on i.line_id=ml.id
        		join finance_approval app on i.approval_id=app.id
        		join finance_visit vis on app.visit_id = vis.id
        		join finance_order ord on vis.order_id = ord.id
        		full join finance_individual_order indiv on indiv.order_id = ord.id 
			full join finance_group_order gord on gord.order_id = ord.id 

        		where i.due_date < %s
        		group by ord.partner_id
        		having sum(ml.amount_residual) > 0 ) s '''

        for col_pos in col_posi:
            self._cr.execute(query, [main_cells_date[i][32:42]])

            query_result = self._cr.fetchall()[0][0]
            if not query_result:
                query_result = 0

            ws[col_pos + str(385)] = query_result
            i += 1
        i = 0



        # Number of Personnel
        # line 389
        # will add it when link HR module

        # Number of Loan Officers
        # will add it when link HR module


        #save seep in current memory
        wb.save(file[:-1])



        #convert seep to decode64 and save it in database and then download it to client

        encoded = 0

        import base64

        with open(file[:-1]) as f:
            encoded = base64.b64encode(f.read())

        attachment_id = self.env['ir.attachment'].create(
            {
                'name': 'SEEP.xls',
                'datas': encoded,
                'datas_fname': 'SEEP.xls',
                'res_model': self._name,
                'type': 'binary'
            })

        return {
            
        'type': 'ir.actions.act_url',
            #'url': 'http://localhost:8069/web/content/?model=ir.attachment&id=%s&filename_field=datas_fname&field=datas&download=true&filename=SEEP-.xls' % attachment_id.id,
                'url': 'http://192.168.43.8:8069/web/content/?model=ir.attachment&id=%s&filename_field=datas_fname&field=datas&download=true&filename=SEEP-.xls' % attachment_id.id,

                'target': 'self',

        }


    ################### And Finally End OF SEEP ^__^ ########################